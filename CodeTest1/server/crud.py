from fastapi import HTTPException
import openpyxl
from fastapi.responses import FileResponse
from sqlalchemy import asc
from sqlalchemy.orm import Session
from models import Part, Time, NewPartData
from datetime import datetime
import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter



def get_times(db: Session, skip: int = 0, limit: int = 10) -> list[Time]:
    return db.query(Time).order_by(Time.start_part_id, Time.end_part_id).offset(skip).limit(limit).all()

def add_new_part(db: Session, new_part_data: NewPartData):
    try:
        # 1. Create new part
        db_part = Part(part_no=new_part_data.new_part_no)
        db.add(db_part)
        db.flush()  # Flush the session to get the part ID without committing

        # 2. Create initial time relation with time=0
        start_part = db.query(Part).filter(Part.part_no == new_part_data.new_part_no).first()
        end_part = db.query(Part).filter(Part.part_no == new_part_data.new_part_no).first()
        
        if start_part and end_part:
            db_time = Time(
                start_part_id=start_part.id,
                end_part_id=end_part.id,
                time=0,
                active=1,
                update_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
            db.add(db_time)

        # 3. Create time relations from part_relations
        for relation in new_part_data.part_relations:
            start_part = db.query(Part).filter(Part.part_no == relation.start_part_no).first()
            end_part = db.query(Part).filter(Part.part_no == relation.end_part_no).first()
            
            if start_part and end_part:
                db_time = Time(
                    start_part_id=start_part.id,
                    end_part_id=end_part.id,
                    time=relation.time,
                    active=1,
                    update_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                db.add(db_time)

        # Commit only after all operations are successful
        db.commit()
        db.refresh(db_part)
        db.refresh(db_time)

        return db_part

    except Exception as e:
        # Rollback in case of error
        db.rollback()
        raise e  # Re-raise the exception to be handled by the calling code

def get_parts_with_times(db: Session):
    parts = db.query(Part).all()
    
    part_to_times = {}

    for part in parts:
        times = db.query(Time).filter(
            (Time.start_part_id == part.id)
        ).order_by(asc(Time.start_part_id), asc(Time.end_part_id)).all()

        part_to_times[part.part_no] = {t.end_part.part_no: t.time for t in times}

    all_part_nos = list(part_to_times.keys())

    headers = [{'text': "Part NO", 'value': 'part_no'}] + [{'text': part_no, 'value': part_no} for part_no in all_part_nos]
    
    rows = []
    for part_no in all_part_nos:
        row = {'part_no': part_no}
        for header in all_part_nos:
            if part_no == header:
                row[header] = "dense"
            else:
                value = part_to_times.get(part_no, {}).get(header, 0)
                row[header] = value if value != 0 else "dense"
        rows.append(row)

    return {
        'headers': headers,
        'rows': rows
    }

def generate_excel_file(db: Session) -> str:
    data = get_parts_with_times(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parts Change over matrix"

    headers = data['headers']
    rows = data['rows']

    # Styles
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    subheader_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    dense_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write title
    title_cell = ws['A1']
    title_cell.value = "Parts Change over matrix"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header['text'])
        cell.fill = header_fill if col_num == 1 else subheader_fill
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write rows
    for row_num, row_data in enumerate(rows, 3):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header['value'], "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if col_num == 1:
                cell.fill = subheader_fill
                cell.font = Font(bold=True)
            elif value == "dense":
                cell.fill = dense_fill
                cell.value = ""
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    # Set a fixed width for all columns except the first one
    fixed_width = 10  # You can adjust this value as needed
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = fixed_width

    # Adjust the width of the first column (title column) based on content
    ws.column_dimensions['A'].width = max(len(str(cell.value)) for cell in ws['A']) + 2

    # Save the file
    file_path = "parts_change_over_matrix.xlsx"  # Save in the current directory
    wb.save(file_path)
    wb.close()

    return file_path

def process_excel_file(db: Session, file_path: str):
    try:
        # Load the Excel file
        wb = load_workbook(file_path)
        ws = wb.active

        # Read data from the Excel sheet
        head = []
        row_detail = []

        row_index = 2  # เริ่มต้นที่แถวที่ 2

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] == 'Part NO':
                head = list(row)
            else:
                row_data = []
                for idx, cell in enumerate(row):
                    if cell is None:
                        cell = 0
                    
                    if idx > 0 and not isinstance(cell, (int, float)):
                        raise ValueError(f"Invalid data type in row {row_index}, column {idx + 1}. Expected int or float, got {type(cell).__name__}.")
                    
                    row_data.append(cell)
                if len(row_data) != len(head):
                    raise ValueError(f"Row length mismatch at row {row_index}. Expected {len(head)}, got {len(row_data)}.")
                
                row_detail.append(row_data)
            
            row_index += 1
        
        if len(row_detail) != len(head) - 1:
            raise ValueError(f"Mismatch between head length minus one ({len(head) - 1}) and the number of data rows ({len(row_detail)}).")

        for index, data in enumerate(row_detail):
            if data[0] != head[index + 1]:
                raise ValueError(f"Mismatch at row_detail index {index}: data[0] ({data[0]}) does not match head[{index + 1}] ({head[index + 1]}).")
            
            # Check for more than one zero value in the data
            zero_count = data.count(0)
            if zero_count > 1:
                raise ValueError(f"More than one zero found in data at row_detail index {index}.")
        
        db.query(Time).update({"active": 0}, synchronize_session=False)
        db.commit() 
        for part_no in head[1:]:  # ข้ามค่า 'Part NO'
            if not db.query(Part).filter_by(part_no=part_no).first():
                db_part = Part(part_no=part_no)
                db.add(db_part)
        db.commit() 
        print("head:", head)
        print("row_detail:", row_detail)
        for index, data in enumerate(row_detail):
            start_part_no = data[0]  # กำหนดค่า start_part_no จากข้อมูลใน data
            for index_hed, data_head in enumerate(head[1:]):
                

                s_part_id = db.query(Part).filter_by(part_no=start_part_no).first()
                if s_part_id is None:
                    raise ValueError(f"Start part_no {start_part_no} not found in the database.")
                
                e_part_id = db.query(Part).filter_by(part_no=data_head).first()
                if e_part_id is None:
                    raise ValueError(f"End part_no {data_head} not found in the database.")

                # Check if a Time entry with the given start_part_id and end_part_id already exists
                existing_time_entry = db.query(Time).filter_by(
                    start_part_id=s_part_id.id,
                    end_part_id=e_part_id.id
                ).first()
                print("เริ่ม",start_part_no)
                print("ถึง",data_head)
                print("ข้อมูล",data)
                """print("แถว",index)
                print("หัว",index_hed)"""
                if existing_time_entry:
                    
                    # Update existing Time entry
                    time_value = data[index_hed + 1] if s_part_id.id != e_part_id.id else 0
                    print("time",time_value)
                    if s_part_id.id == e_part_id.id:
                        """print("1.1")
                        print(f"Start Part ID: {s_part_id.id}, End Part ID: {e_part_id.id}, Existing Time Entry: {existing_time_entry}")
                        print(f"Found existing entry: {existing_time_entry.id}, Time: {existing_time_entry.time}")"""
                        existing_time_entry.active = 1
                        existing_time_entry.update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        """print("1.2")
                        print(f"Start Part ID: {s_part_id.id}, End Part ID: {e_part_id.id}, Existing Time Entry: {existing_time_entry}")
                        print(f"Found existing entry: {existing_time_entry.id}, Time: {existing_time_entry.time}")"""
                        existing_time_entry.time = time_value
                        existing_time_entry.active = 1
                        existing_time_entry.update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                else:
                    # Create a new Time entry
                    new_time = Time(
                        start_part_id=s_part_id.id,
                        end_part_id=e_part_id.id,
                        time=data[index_hed + 1] if s_part_id.id != e_part_id.id else 0,
                        active=1,
                        update_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    )
                    db.add(new_time)
        db.commit() 
        db.query(Time).filter_by(active=0).delete(synchronize_session=False)
        part_numbers_in_db = db.query(Part).all()
        for part in part_numbers_in_db:
            if part.part_no not in head:
                db.delete(part)
        db.commit()
        
        print("head:", head)
        print("row_detail:", row_detail)

        return file_path

    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def delete_part(db: Session, part_no: str):
    part = db.query(Part).filter(Part.part_no == part_no).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    db.query(Time).filter(
        (Time.start_part_id == part.id) | (Time.end_part_id == part.id)
    ).delete(synchronize_session=False)
    
    db.delete(part)
    db.commit()
    return {"message": "Part deleted successfully"}